from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side


def format_excel_datetime(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        if value.tzinfo is not None:
            value = value.replace(tzinfo=None)
        return value.strftime("%d.%m.%Y %H:%M")
    return str(value)


def format_excel_date(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y")
    if hasattr(value, "strftime"):
        return value.strftime("%d.%m.%Y")
    return str(value)


def safe_enum_value(value):
    if value is None:
        return ""
    return value.value if hasattr(value, "value") else str(value)


def get_material_name(item):
    if getattr(item, "is_manual", False):
        return getattr(item, "manual_name", "") or ""

    agreement_material = getattr(item, "agreement_material", None)
    if agreement_material:
        return getattr(agreement_material, "name", "") or ""

    return ""


def get_material_unit(item):
    if getattr(item, "is_manual", False):
        return safe_enum_value(getattr(item, "manual_unit", ""))

    agreement_material = getattr(item, "agreement_material", None)
    if agreement_material:
        return safe_enum_value(getattr(agreement_material, "unit", ""))

    return ""


def get_material_characteristics(item):
    if getattr(item, "is_manual", False):
        return getattr(item, "manual_comment", "") or ""

    agreement_material = getattr(item, "agreement_material", None)
    if agreement_material:
        for attr_name in ("characteristics", "description", "specification"):
            value = getattr(agreement_material, attr_name, None)
            if value not in (None, ""):
                return str(value)

    return ""


def generate_request_excel(request):
    print("=== GENERATE REQUEST EXCEL START ===")
    print("REQUEST ID:", getattr(request, "id", None))
    print("REQUEST OBJECT:", safe_enum_value(getattr(request, "object", None)))
    print("REQUEST AGREEMENT:", getattr(request, "agreement", None))
    print("REQUEST SECTION:", getattr(request, "section", None))
    print("REQUEST DELIVERY DATE:", getattr(request, "delivery_date", None))

    materials = list(getattr(request, "materials", []) or [])
    print("MATERIALS COUNT:", len(materials))

    wb = Workbook()
    ws = wb.active
    ws.title = "Заявка"

    title_font = Font(bold=True, size=14)
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 24
    ws.column_dimensions["G"].width = 24

    # Гарантированно заметный текст
    ws.merge_cells("A1:G1")
    ws["A1"] = "Форма заявки на поставку материалов"
    ws["A1"].font = title_font
    ws["A1"].alignment = center

    ws["A4"] = "Номер заявки:"
    ws["B4"] = getattr(request, "id", "")

    ws["A5"] = "Наименование:"
    ws["B5"] = getattr(request, "name", "") or ""

    ws["A6"] = "Объект:"
    ws["B6"] = safe_enum_value(getattr(request, "object", ""))

    ws["A7"] = "Секция:"
    ws["B7"] = getattr(request, "section", "") or ""

    ws["A8"] = "Примечание:"
    ws["B8"] = getattr(request, "description", "") or ""

    ws["A9"] = "Шифр проекта:"
    ws["B9"] = getattr(request, "agreement", "") or ""

    ws["A10"] = "Дата поставки:"
    ws["B10"] = format_excel_date(getattr(request, "delivery_date", None))

    ws["A11"] = "Дата создания:"
    ws["B11"] = format_excel_datetime(getattr(request, "created_at", None))

    for cell in ("A4", "A5", "A6", "A7", "A8", "A9", "A10", "A11"):
        ws[cell].font = bold

    start_row = 13
    headers = [
        "№",
        "Наименование материала",
        "Комментарий",
        "Ед. изм.",
        "Кол-во",
        "Примечание",
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.font = bold
        cell.alignment = center
        cell.border = border

    if not materials:
        row = start_row + 1
        for col in range(1, 8):
            c = ws.cell(row=row, column=col, value="")
            c.border = border
        ws.cell(row=row, column=2, value="Материалы отсутствуют")
    else:
        for i, item in enumerate(materials, start=1):
            row = start_row + i
            notes = []
            if getattr(item, "is_manual", False):
                notes.append("Введено вручную")
            if getattr(item, "overdraft", False):
                notes.append("Перерасход")

            ws.cell(row=row, column=1, value=i).border = border
            ws.cell(row=row, column=2, value=get_material_name(item)).border = border
            ws.cell(row=row, column=3, value=get_material_characteristics(item)).border = border
            ws.cell(row=row, column=4, value=get_material_unit(item)).border = border
            ws.cell(row=row, column=5, value=float(getattr(item, "quantity", 0) or 0)).border = border
            ws.cell(row=row, column=6, value="; ".join(notes)).border = border

    sign_row = start_row + max(len(materials), 1) + 3
    ws["A" + str(sign_row)] = "Составил:"
    ws["B" + str(sign_row)] = getattr(request, "author_name", "") or ""
    ws["E" + str(sign_row)] = "Дата:"
    ws["F" + str(sign_row)] = format_excel_date(date.today())

    ws.merge_cells(f"A{sign_row + 1}:B{sign_row + 1}")
    ws[f"A{sign_row + 1}"] = "Согласовано: Руководитель проекта"
    OBJECT_TO_NAME = {
        "AURIKA": "Ханов Ильшат Милхатович",
        "AURUM": "Асадуллин Айнур Ильшатович",
        "MAXIMUS": "Родионов Андрей Николаевич",
    }

    obj = getattr(request, "object", None)
    fio = OBJECT_TO_NAME.get(obj.name, "") if obj else ""

    ws.cell(row=sign_row + 1, column=3, value=fio)

    output_dir = Path("generated")
    output_dir.mkdir(parents=True, exist_ok=True)

    file_path = output_dir / f"request_{request.id}.xlsx"
    wb.save(file_path)
    wb.close()

    return str(file_path.resolve())
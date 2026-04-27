from pathlib import Path
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook

from backend.database import SessionLocal
from backend.models.enum import ObjectsEnum
from backend.models.agreement import AgreementMaterial


EXCEL_PATH = Path(__file__).resolve().parents[1] / "data" / "materials.xlsx"

OBJECT = ObjectsEnum.AURUM

SHEET_NAME = "Заявка"

START_ROW = 3


def to_decimal(value):
    if value is None or value == "":
        return None

    try:
        return Decimal(str(value).replace(",", "."))
    except InvalidOperation:
        return None


def normalize_text(value):
    if value is None:
        return ""

    return str(value).strip()


def import_materials():
    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb[SHEET_NAME]

    db = SessionLocal()

    created = 0
    updated = 0
    skipped = 0

    try:
        for row in range(START_ROW, ws.max_row + 1):
            name = normalize_text(ws.cell(row=row, column=1).value)
            unit = normalize_text(ws.cell(row=row, column=2).value)
            quantity = to_decimal(ws.cell(row=row, column=3).value)

            if not name or not unit or quantity is None:
                skipped += 1
                continue

            existing = (
                db.query(AgreementMaterial)
                .filter(
                    AgreementMaterial.name == name,
                    AgreementMaterial.unit == unit,
                    AgreementMaterial.object == OBJECT,
                )
                .first()
            )

            if existing:
                existing.total_quantity = quantity
                updated += 1
            else:
                material = AgreementMaterial(
                    name=name,
                    unit=unit,
                    object=OBJECT,
                    total_quantity=quantity,
                    reserved_quantity=0,
                    spent_quantity=0,
                )
                db.add(material)
                created += 1

        db.commit()

        print(f"Импорт завершён")
        print(f"Создано: {created}")
        print(f"Обновлено: {updated}")
        print(f"Пропущено: {skipped}")

    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


if __name__ == "__main__":
    import_materials()